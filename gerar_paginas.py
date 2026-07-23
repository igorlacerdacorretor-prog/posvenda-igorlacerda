#!/usr/bin/env python3
"""
Gera as páginas de acompanhamento de pós-venda a partir da planilha Excel.

Uso:
    python gerar_paginas.py "Pós Venda 2026.xlsx"
    python gerar_paginas.py "Pós Venda 2026.xlsx" --aba "Pós Venda 2026"

O que o script faz:
    1. Lê a aba de pós-venda do Excel.
    2. Para cada venda, gera (ou reaproveita) um código único e aleatório
       que nunca muda entre execuções (guardado em codigos_clientes.csv).
    3. Monta um JSON por cliente em site/data/<codigo>.json com as etapas
       do processo (Financiamento ou À vista), status de cada uma
       (pendente / andamento / concluído) e progresso geral.
    4. Gera links_para_enviar.csv com o link pronto de cada cliente para
       você copiar e colar no WhatsApp.

Depois de rodar, é só subir as pastas "site/" (com o data/ atualizado)
para o GitHub Pages -- veja o README.md para o passo a passo.
"""
from __future__ import annotations

import argparse
import csv
import json
import random
import string
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
CODIGOS_PATH = BASE_DIR / "codigos_clientes.csv"
CONFIG_PATH = BASE_DIR / "config.json"
LINKS_PATH = BASE_DIR / "links_para_enviar.csv"

DEFAULT_ABA = "Pós Venda 2026"
PRIMEIRA_LINHA_DADOS = 3  # linha 1 = título dos grupos, linha 2 = cabeçalho, linha 3 = 1ª venda

# Colunas fixas (A a F) -----------------------------------------------------
COL_NUM_VENDA = "A"
COL_FORMA_PAGAMENTO = "B"
COL_NOME = "C"
COL_TIPO = "D"
COL_DATA_CONTRATO = "E"
COL_STATUS = "F"

# Etapas do fluxo "Financiamento": colunas G-O e Q-X -------------------------
# (coluna, nome original na planilha, nome amigável exibido ao cliente)
ETAPAS_FINANCIAMENTO = [
    ("G", "Envio doc. Compradores", "Envio de documentos do comprador"),
    ("H", "Aprov. do crédito", "Aprovação do crédito"),
    ("I", "Envio doc. Vendedores", "Envio de documentos do vendedor"),
    ("J", "Vistoria Engenheiro", "Vistoria do engenheiro"),
    ("K", "Aprov. Laudo Engenharia", "Aprovação do laudo de engenharia"),
    ("L", "Ent. Corresp. Caixa", "Entrega ao correspondente bancário"),
    ("M", "Doc. Conformidade", "Conformidade da documentação"),
    ("N", "Ent. Gerencial", "Entrevista gerencial"),
    ("O", "Ass. Contrato", "Assinatura do contrato"),
    ("Q", "Envio Sol. ITBI", "Solicitação do ITBI"),
    ("R", "Pagamento", "Pagamento do ITBI"),
    ("S", "CND ITBI", "Certidão de quitação do ITBI"),
    ("T", "Prenotação Cartório", "Prenotação em cartório"),
    ("U", "Pagamento Registro", "Pagamento do registro"),
    ("V", "Escritura Registrada", "Registro do imóvel concluído"),
    ("W", "Dev. Contrato", "Devolução do contrato ao banco"),
    ("X", "Pagamento Vendedor", "Pagamento ao vendedor"),
]

# Etapas do fluxo "À vista": colunas Z-AG ------------------------------------
ETAPAS_A_VISTA = [
    ("Z", "Envio Sol. ITBI", "Solicitação do ITBI"),
    ("AA", "Pagamento", "Pagamento do ITBI"),
    ("AB", "CND ITBI", "Certidão de quitação do ITBI"),
    ("AC", "Ent. Cart. Notas - Lavratura", "Entrega ao cartório de notas"),
    ("AD", "Escritura Lavratura", "Lavratura da escritura"),
    ("AE", "Prenotação Cartório", "Prenotação em cartório"),
    ("AF", "Pagamento Registro", "Pagamento do registro"),
    ("AG", "Escritura Registrada", "Registro do imóvel concluído"),
]

CODIGO_ALFABETO = string.ascii_lowercase + string.digits


def carregar_config() -> dict:
    if CONFIG_PATH.exists():
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    config = {"base_url": "https://SEU-USUARIO.github.io/SEU-REPOSITORIO/pagina.html"}
    CONFIG_PATH.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Criei {CONFIG_PATH.name} com uma URL de exemplo. Edite o 'base_url' com o endereço real do seu GitHub Pages.")
    return config


def carregar_codigos() -> dict:
    codigos = {}
    if CODIGOS_PATH.exists():
        with open(CODIGOS_PATH, newline="", encoding="utf-8") as f:
            for linha in csv.DictReader(f):
                codigos[linha["numero_venda"]] = linha["codigo"]
    return codigos


def salvar_codigos(codigos: dict) -> None:
    with open(CODIGOS_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["numero_venda", "codigo"])
        writer.writeheader()
        for numero_venda, codigo in codigos.items():
            writer.writerow({"numero_venda": numero_venda, "codigo": codigo})


def gerar_codigo(existentes: set) -> str:
    while True:
        codigo = "".join(random.choices(CODIGO_ALFABETO, k=8))
        if codigo not in existentes:
            return codigo


def valor_data(valor) -> str | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def montar_etapas(ws, numero_linha: int, definicoes: list[tuple[str, str, str]]) -> list[dict]:
    valores = []
    for coluna, _nome_original, nome_amigavel in definicoes:
        valor = ws[f"{coluna}{numero_linha}"].value
        valores.append((nome_amigavel, valor))

    ultima_concluida_idx = -1
    for idx, (_nome, valor) in enumerate(valores):
        if valor not in (None, ""):
            ultima_concluida_idx = idx

    etapas = []
    for idx, (nome_amigavel, valor) in enumerate(valores):
        concluida = valor not in (None, "")
        if concluida:
            status = "concluido"
        elif idx == ultima_concluida_idx + 1:
            status = "andamento"
        else:
            status = "pendente"
        etapas.append(
            {
                "nome": nome_amigavel,
                "status": status,
                "data": valor_data(valor) if concluida else None,
                "observacao": None,
            }
        )
    return etapas


def calcular_progresso(etapas: list[dict]) -> int:
    if not etapas:
        return 0
    concluidas = sum(1 for e in etapas if e["status"] == "concluido")
    return round(100 * concluidas / len(etapas))


def processar(caminho_excel: str, aba: str) -> None:
    wb = load_workbook(caminho_excel, data_only=True)
    if aba not in wb.sheetnames:
        print(f"ERRO: a aba '{aba}' não existe nesse arquivo.")
        print(f"Abas disponíveis: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    ws = wb[aba]

    config = carregar_config()
    codigos = carregar_codigos()
    codigos_em_uso = set(codigos.values())

    pasta_dados = BASE_DIR / "site" / "data"
    pasta_dados.mkdir(parents=True, exist_ok=True)

    linhas_geradas = []

    for numero_linha in range(PRIMEIRA_LINHA_DADOS, ws.max_row + 1):
        numero_venda = ws[f"{COL_NUM_VENDA}{numero_linha}"].value
        nome = ws[f"{COL_NOME}{numero_linha}"].value
        if numero_venda is None or not nome:
            continue

        forma_pagamento = str(ws[f"{COL_FORMA_PAGAMENTO}{numero_linha}"].value or "").strip()
        tipo = ws[f"{COL_TIPO}{numero_linha}"].value
        data_contrato = ws[f"{COL_DATA_CONTRATO}{numero_linha}"].value
        status_geral = str(ws[f"{COL_STATUS}{numero_linha}"].value or "").strip()

        numero_venda_str = str(numero_venda)
        if numero_venda_str not in codigos:
            novo_codigo = gerar_codigo(codigos_em_uso)
            codigos[numero_venda_str] = novo_codigo
            codigos_em_uso.add(novo_codigo)
        codigo = codigos[numero_venda_str]

        definicoes = ETAPAS_A_VISTA if forma_pagamento.lower() == "à vista" else ETAPAS_FINANCIAMENTO
        etapas = montar_etapas(ws, numero_linha, definicoes)

        cancelado = status_geral.lower() == "cancelado"
        if cancelado:
            for etapa in etapas:
                if etapa["status"] == "andamento":
                    etapa["status"] = "pendente"

        dados_cliente = {
            "codigo": codigo,
            "cliente": str(nome),
            "imovel": str(tipo) if tipo else None,
            "forma_pagamento": forma_pagamento or None,
            "status_geral": status_geral or None,
            "cancelado": cancelado,
            "data_contrato": valor_data(data_contrato),
            "progresso_pct": calcular_progresso(etapas),
            "etapas": etapas,
            "atualizado_em": date.today().strftime("%d/%m/%Y"),
        }

        caminho_json = pasta_dados / f"{codigo}.json"
        caminho_json.write_text(
            json.dumps(dados_cliente, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        link = f"{config['base_url']}?id={codigo}"
        linhas_geradas.append(
            {"numero_venda": numero_venda_str, "cliente": nome, "codigo": codigo, "link": link}
        )

    salvar_codigos(codigos)

    with open(LINKS_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["numero_venda", "cliente", "codigo", "link"])
        writer.writeheader()
        writer.writerows(linhas_geradas)

    print(f"\n{len(linhas_geradas)} página(s) de cliente geradas em site/data/")
    print(f"Links prontos para enviar: {LINKS_PATH.name}\n")
    for linha in linhas_geradas:
        print(f"  {linha['cliente']}: {linha['link']}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera as páginas de pós-venda a partir do Excel.")
    parser.add_argument("planilha", help="Caminho do arquivo .xlsx exportado do CRM")
    parser.add_argument("--aba", default=DEFAULT_ABA, help=f"Nome da aba (padrão: '{DEFAULT_ABA}')")
    args = parser.parse_args()

    if not Path(args.planilha).exists():
        print(f"ERRO: arquivo não encontrado: {args.planilha}")
        sys.exit(1)

    processar(args.planilha, args.aba)


if __name__ == "__main__":
    main()
